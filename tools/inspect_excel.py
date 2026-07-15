import openpyxl
import os

mom_excel = r"H:\Program Files(x86)\Activision\Call To Power 2\Scenarios\mom\mom_dimension_inventory.xlsx"
ae_excel = r"H:\Program Files(x86)\Activision\Call To Power 2\Scenarios\AE_Mod\ae_mod_control_plane_v2.xlsx"

print("=== MoM Dimension Inventory ===")
if os.path.exists(mom_excel):
    wb_mom = openpyxl.load_workbook(mom_excel, data_only=True)
    print("Sheets:", wb_mom.sheetnames)
    for sheet_name in wb_mom.sheetnames[:5]: 
        ws = wb_mom[sheet_name]
        headers = [cell.value for cell in ws[1]]
        row2 = [cell.value for cell in ws[2]]
        print(f"\n[Sheet: {sheet_name}]")
        print("Headers:", headers[:15])
        print("Sample Row:", row2[:15])
else:
    print("MoM Excel not found.")

print("\n=== AE Mod Control Plane v2 ===")
if os.path.exists(ae_excel):
    wb_ae = openpyxl.load_workbook(ae_excel, data_only=True)
    print("Sheets:", wb_ae.sheetnames)
    for sheet_name in wb_ae.sheetnames[:5]:
        ws = wb_ae[sheet_name]
        headers = [cell.value for cell in ws[1]]
        row2 = [cell.value for cell in ws[2]]
        print(f"\n[Sheet: {sheet_name}]")
        print("Headers:", headers[:15])
        print("Sample Row:", row2[:15])
else:
    print("AE Excel not found.")