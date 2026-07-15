import openpyxl
import os

# Paths
ae_excel_path = r"H:\Program Files(x86)\Activision\Call To Power 2\Scenarios\AE_Mod\ae_mod_control_plane_v2.xlsx"
output_excel_path = r"H:\Program Files(x86)\Activision\Call To Power 2\Scenarios\AE_Mod\ae_mod_control_plane_v2_UPDATED.xlsx"

# Age cutoff logic: Allow up to Industrial/Renaissance. Exclude Future/Modern.
# Adjust this list based on the exact age names in your AE mod.
ALLOWED_AGES = {"AGE_ONE", "AGE_TWO", "AGE_THREE", "AGE_FOUR", "AGE_FIVE"}
EXCLUDED_AGES = {"AGE_SIX"}

SHEETS_TO_UPDATE = ["advances", "units", "improvements", "wonders"]

def update_workbook():
    if not os.path.exists(ae_excel_path):
        print(f"ERROR: File not found at {ae_excel_path}")
        return

    print(f"Loading {ae_excel_path}...")
    wb = openpyxl.load_workbook(ae_excel_path)
    
    # First, build a lookup of advance_id -> age from the 'advances' sheet
    # so we can infer the age for units/improvements/wonders based on their prereq.
    advance_age_map = {}
    if "advances" in wb.sheetnames:
        ws_adv = wb["advances"]
        headers = [cell.value for cell in ws_adv[1]]
        try:
            id_idx = headers.index("advance_id")
            age_idx = headers.index("age")
            
            for row in ws_adv.iter_rows(min_row=2, values_only=True):
                if row[id_idx]:
                    advance_age_map[str(row[id_idx]).strip().upper()] = str(row[age_idx]).strip().upper() if row[age_idx] else ""
        except ValueError:
            print("Warning: 'advance_id' or 'age' column not found in advances sheet.")

    for sheet_name in SHEETS_TO_UPDATE:
        if sheet_name not in wb.sheetnames:
            print(f"Skipping sheet '{sheet_name}': not found.")
            continue
            
        print(f"\nProcessing sheet: {sheet_name}")
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        # Determine ID and Prereq column indices (they vary by sheet)
        id_col_name = "advance_id" if sheet_name == "advances" else "name"
        prereq_col_name = "prereq" if sheet_name in ["units", "improvements"] else "enable_advance"
        
        try:
            id_idx = headers.index(id_col_name)
        except ValueError:
            print(f"  Warning: '{id_col_name}' not found in {sheet_name}. Skipping age inference.")
            id_idx = -1
            
        try:
            prereq_idx = headers.index(prereq_col_name)
        except ValueError:
            prereq_idx = -1

        # Add new columns if they don't exist
        if "SourceMod" not in headers:
            ws.cell(row=1, column=len(headers) + 1, value="SourceMod")
            source_mod_idx = len(headers) + 1
        else:
            source_mod_idx = headers.index("SourceMod") + 1
            
        if "IncludeInMoM" not in headers:
            ws.cell(row=1, column=len(headers) + 2, value="IncludeInMoM")
            include_idx = len(headers) + 2
        else:
            include_idx = headers.index("IncludeInMoM") + 1
            
        # For advances, also ensure 'age' exists (it should)
        has_age_col = "age" in headers

        # Process data rows
        modified_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Set SourceMod
            ws.cell(row=row_idx, column=source_mod_idx, value="AE")
            
            # Determine IncludeInMoM
            include_val = True # Default to True
            
            if sheet_name == "advances" and has_age_col:
                age_idx = headers.index("age")
                age_val = str(row[age_idx].value).strip().upper() if row[age_idx].value else ""
                if age_val in EXCLUDED_AGES:
                    include_val = False
            else:
                # For units/improvements/wonders, try to infer from prereq/enable_advance
                if prereq_idx != -1 and id_idx != -1:
                    prereq_val = str(row[prereq_idx].value).strip().upper() if row[prereq_idx].value else ""
                    # Prereq can be a single advance or semicolon-separated list
                    prereqs = [p.strip() for p in prereq_val.split(";")]
                    for p in prereqs:
                        if p in advance_age_map and advance_age_map[p] in EXCLUDED_AGES:
                            include_val = False
                            break
            
            ws.cell(row=row_idx, column=include_idx, value=include_val)
            if not include_val:
                modified_count += 1
                
        print(f"  Updated {modified_count} records to IncludeInMoM=False based on age/prereq.")

    print(f"\nSaving updated workbook to: {output_excel_path}")
    wb.save(output_excel_path)
    print("DONE! Please review the new file in Excel before replacing the original.")

if __name__ == "__main__":
    update_workbook()
