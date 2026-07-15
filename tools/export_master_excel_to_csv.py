import openpyxl
import csv
import os
from pathlib import Path

MASTER_EXCEL = r"H:\Program Files(x86)\Activision\Call To Power 2\Scenarios\AE_Mod\ae_mod_control_plane_v2_UPDATED.xlsx"
GENERATOR_CSV_DIR = Path(r"H:\Program Files(x86)\Activision\Call To Power 2\Scenarios\mom\tools\momjr_csv")

# Hardcoded blocklist of known future-age / out-of-genre AE improvements to force Exclude
OUT_OF_GENRE_IMPROVEMENTS = {
    "IMPROVE_BEHAVIORAL_MOD_CENTER", "IMPROVE_NUCLEAR_PLANT", "IMPROVE_ROBOTIC_PLANT",
    "IMPROVE_GAIA_COMPUTER", "IMPROVE_INCUBATION_CENTER", "IMPROVE_MATTER_DECOMPILER",
    "IMPROVE_MICRO_DEFENSE", "IMPROVE_ORBITAL_LABORATORY", "IMPROVE_POWER_SATELLITE",
    "IMPROVE_PUBLIC_TRANSPORTATION", "IMPROVE_PUBLISHING_HOUSE", "IMPROVE_SECURITY_MONITOR",
    "IMPROVE_VR_AMUSEMENT_PARK", "IMPROVE_WATER_TRANSFORMER", "IMPROVE_ZERO_EMISSION_VEHICLES",
    "IMPROVE_ARTIFICIAL_WOMB_COMPLEX", "IMPROVE_RAIL_LAUNCHER"
}

def export_to_csv():
    if not os.path.exists(MASTER_EXCEL):
        print(f"ERROR: Master Excel file not found at: {MASTER_EXCEL}")
        return

    print(f"Loading master workbook: {MASTER_EXCEL}")
    wb = openpyxl.load_workbook(MASTER_EXCEL, data_only=True)
    GENERATOR_CSV_DIR.mkdir(parents=True, exist_ok=True)

    # Export improvements
    if "improvements" in wb.sheetnames:
        print("Exporting: improvements")
        ws = wb["improvements"]
        csv_path = GENERATOR_CSV_DIR / "improvements.csv"
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            headers = ['name', 'description', 'icon', 'upkeep', 'prereq', 'enable_advance', 'production_cost', 'SourceMod', 'IncludeInMoM']
            writer.writerow(headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1]: continue # skip empty rows (col 1 is 'name' in AE schema)
                name = str(row[1]) if row[1] else ""
                desc = str(row[2]) if row[2] else ""
                icon = str(row[3]) if row[3] else ""
                upkeep = str(row[4]) if row[4] else "0"
                cost = str(row[7]) if row[7] else "100"
                
                # Force blocklist
                if name.upper() in OUT_OF_GENRE_IMPROVEMENTS:
                    include = "False"
                else:
                    include = str(row[-1]) if len(row) > 1 and row[-1] else "True"
                    
                writer.writerow([name, desc, icon, upkeep, "", "", cost, "AE", include])

    # Export advances (reuse previous logic but ensure we read the IncludeInMoM column correctly)
    if "advances" in wb.sheetnames:
        print("Exporting: advances")
        ws = wb["advances"]
        csv_path = GENERATOR_CSV_DIR / "advances.csv"
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['name', 'epoch', 'category', 'prereq1', 'prereq2', 'cost', 'SourceMod', 'IncludeInMoM'])
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                name = str(row[0]) if row[0] else ""
                cost = str(row[1]) if row[1] else "1"
                age = str(row[2]) if row[2] else ""
                prereqs = str(row[4]) if row[4] else ""
                include = str(row[-1]) if len(row) > 1 and row[-1] else "True"
                
                epoch_map = {"AGE_ONE": "1", "AGE_TWO": "2", "AGE_THREE": "3", "AGE_FOUR": "4", "AGE_FIVE": "5", "AGE_SIX": "6"}
                epoch = epoch_map.get(age.strip().upper(), age)
                prereq_list = [p.strip() for p in prereqs.split(';') if p.strip()]
                p1 = prereq_list[0] if len(prereq_list) > 0 else ""
                p2 = prereq_list[1] if len(prereq_list) > 1 else ""
                
                writer.writerow([name, epoch, "", p1, p2, cost, "AE", include])

    # Export units
    if "units" in wb.sheetnames:
        print("Exporting: units")
        ws = wb["units"]
        csv_path = GENERATOR_CSV_DIR / "units.csv"
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            headers = ['name', 'description', 'attack', 'defense', 'hp', 'firepower', 'prereq', 'cost', 'move', 'SourceMod', 'IncludeInMoM']
            writer.writerow(headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                name = str(row[0]) if row[0] else ""
                desc = str(row[1]) if row[1] else ""
                atk = str(row[2]) if row[2] else "0"
                def_ = str(row[3]) if row[3] else "0"
                hp = str(row[4]) if row[4] else "10"
                fp = str(row[5]) if row[5] else "1"
                prereq = str(row[6]) if row[6] else ""
                cost = str(row[11]) if row[11] else "100"
                move = str(row[16]) if len(row) > 16 and row[16] else "1"
                include = str(row[-1]) if len(row) > 1 and row[-1] else "True"
                writer.writerow([name, desc, atk, def_, hp, fp, prereq, cost, move, "AE", include])

    # Export wonders
    if "wonders" in wb.sheetnames:
        print("Exporting: wonders")
        ws = wb["wonders"]
        csv_path = GENERATOR_CSV_DIR / "wonders.csv"
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            headers = ['name', 'description', 'icon', 'enable_advance', 'production_cost', 'SourceMod', 'IncludeInMoM']
            writer.writerow(headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1]: continue
                name = str(row[1]) if row[1] else ""
                desc = str(row[3]) if row[3] else ""
                icon = str(row[7]) if row[7] else ""
                enable = str(row[8]) if row[8] else ""
                cost = str(row[9]) if row[9] else "100"
                include = str(row[-1]) if len(row) > 1 and row[-1] else "True"
                writer.writerow([name, desc, icon, enable, cost, "AE", include])

    print("SUCCESS: Control plane exported to CSVs with MoM-compatible schema and strict genre filtering.")

if __name__ == "__main__":
    export_to_csv()