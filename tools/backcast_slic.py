"""Backcast the scenario's SLIC into the control-plane workbook as the `slic` tab.

WHY THIS EXISTS
---------------
The control plane is `mom_dimension_inventory.xlsx`, and every tab in it is one
DIMENSION of the mod. SLIC is a dimension like any other -- it is simply the one
that cannot be forward-generated, because Civ2 has no equivalent to it and there
is nothing upstream to encode FROM. SLIC is authored directly against CTP2.

    Civ2 RULES.TXT ---encode--> [ mom_dimension_inventory.xlsx ] ---generate--> scenario
                                              ^
    scenario *.slc  ---backcast---------------+

So the workbook stays the single artifact the mod is based on, while the `.slc`
files stay the source of truth for behaviour. This tool NEVER writes SLIC. A
spreadsheet that could regenerate SLIC would be strictly worse than text files
that are diffable, commentable and compilable.

THE CELL CONTRACT
-----------------
A cell is a file, and/or a set of constants, classes and/or functions. Not a
description of one -- the actual text. `slic` has one row per module, and its
content columns carry real source:

    constants   module-scope declarations (the per-player arrays)
    functions   every int_f / void_f, whole
    handlers    every HandleEvent block, whole
    triggers    every UI trigger block, whole
    segments    every alertbox / messagebox block, whole
    source      the entire file, verbatim

That makes the tab editable and reviewable in Excel on its own terms, and it is
what lets the second process work: an LLM reads the dimension, proposes a
feature, writes intent into `purpose`, and the implementation lands in SLIC.

`slic_index` is the flat companion manifest -- one row per declaration, with
signature and load phase -- for when you want to scan or filter rather than read.

WHAT IS DERIVED VS WHAT IS CURATED
----------------------------------
Derived every run: module, phase, include order, and all content columns. These
are re-extracted from the code and overwrite unconditionally.

Curated and merged forward, keyed by name: `purpose` and `status`. Prose written
by a human or an LLM survives regeneration. A declaration added in code arrives
with an empty `purpose` -- the visible TODO. A row whose declaration no longer
exists is dropped and reported.

ORDERING NOTE
-------------
`export_mod_workbook.py` rebuilds the workbook from the control-plane CSVs and
does not know about SLIC. Run this AFTER it:

    python tools/export_mod_workbook.py
    python tools/backcast_slic.py

`--check` exits non-zero when the tab has drifted from the code, so it can gate
CI without writing anything.
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

TOOLS_DIR = Path(__file__).resolve().parent
REPO_DIR = TOOLS_DIR.parent
SCEN_GAMEDATA = REPO_DIR / "scen0000" / "default" / "gamedata"
WORKBOOK = REPO_DIR / "mom_dimension_inventory.xlsx"
INDEX_CSV = TOOLS_DIR / "momjr_csv" / "slic_inventory.csv"
CURATION = TOOLS_DIR / "momjr_csv" / "slic_purpose.json"

SHEET_DIM = "slic"
SHEET_INDEX = "slic_index"

# Excel's hard per-cell ceiling. Nothing we ship is close (largest module is
# ~16 KB) but a silent truncation would be a data-loss bug, so it is explicit.
CELL_MAX = 32767

STOCK_MODULES = {"tut2_main.slc"}

DIM_HEADER = [
    "module", "phase", "include_order", "purpose",
    "constants", "functions", "handlers", "triggers", "segments",
    "n_decls", "bytes", "source",
]
IDX_HEADER = ["module", "kind", "name", "signature", "purpose", "phase", "status"]

PHASE_BY_MODULE = {
    "scenario.slc": "A",
    "tutorial.slc": "A",
    "mom_func.slc": "B",
    "mom_turns.slc": "B",
    "mom_city_effects.slc": "B",
    "mom_msg.slc": "C",
    "mom_magic.slc": "M",
    "mom_spells.slc": "M",
}

RE_HANDLER = re.compile(r"^\s*HandleEvent\(\s*(?P<event>\w+)\s*\)\s*'(?P<name>\w+)'\s*(?P<when>pre|post)?")
RE_FUNC = re.compile(r"^\s*(?P<ret>int_f|void_f)\s+(?P<name>\w+)\s*\((?P<args>[^)]*)\)")
RE_TRIGGER = re.compile(r"^\s*trigger\s+'(?P<name>\w+)'\s+on\s+\"(?P<on>[^\"]*)\"")
RE_SEGMENT = re.compile(r"^\s*(?P<kind>alertbox|messagebox)\s+'(?P<name>\w+)'")
RE_GLOBAL = re.compile(r"^(?:int_t|string_t|unit_t|city_t|location_t|player_t)\s+\w+")


def clip(text: str) -> str:
    if len(text) <= CELL_MAX:
        return text
    keep = CELL_MAX - 80
    return text[:keep] + f"\n... TRUNCATED, {len(text) - keep} chars omitted; read the .slc"


def block_at(lines: list[str], start: int) -> tuple[str, int]:
    """Capture a declaration plus its balanced-brace body.

    Returns the text and the index of the last consumed line. A declaration with
    no brace (a bare forward declaration) yields just its own line.
    """
    depth = 0
    seen = False
    for i in range(start, len(lines)):
        # Strip line comments so a brace inside `// ...` cannot unbalance us.
        code = lines[i].split("//", 1)[0]
        depth += code.count("{") - code.count("}")
        if "{" in code:
            seen = True
        if seen and depth <= 0:
            return "\n".join(lines[start:i + 1]), i
        if not seen and code.strip().endswith(";"):
            return lines[start], i
    return "\n".join(lines[start:]), len(lines) - 1


def leading_comment(lines: list[str], start: int) -> str:
    """The comment block immediately above a declaration, if any."""
    out: list[str] = []
    i = start - 1
    while i >= 0 and lines[i].strip().startswith("//"):
        out.append(lines[i])
        i -= 1
    return "\n".join(reversed(out))


def module_thesis(lines: list[str]) -> str:
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        return stripped.lstrip("/").strip() if stripped.startswith("//") else ""
    return ""


def include_order(scenario: Path) -> dict[str, int]:
    """Load order, read from scenario.slc's #include list -- not hardcoded."""
    order: dict[str, int] = {"scenario.slc": 0}
    if not scenario.is_file():
        return order
    n = 0
    for line in scenario.read_text(encoding="utf-8", errors="replace").splitlines():
        if line.strip().startswith("//"):
            continue
        match = re.search(r'#include\s+"([^"]+)"', line)
        if match:
            n += 1
            order[match.group(1)] = n
    return order


def scan(path: Path) -> dict:
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    module = path.name
    buckets: dict[str, list[str]] = {
        "constants": [], "functions": [], "handlers": [], "triggers": [], "segments": [],
    }
    decls: list[dict[str, str]] = []

    i = 0
    while i < len(lines):
        line = lines[i]
        consumed = i

        if RE_GLOBAL.match(line):
            buckets["constants"].append(line)
            i += 1
            continue

        for regex, bucket, kind in (
            (RE_HANDLER, "handlers", "handler"),
            (RE_FUNC, "functions", "function"),
            (RE_TRIGGER, "triggers", "trigger"),
            (RE_SEGMENT, "segments", None),
        ):
            match = regex.match(line)
            if not match:
                continue
            groups = match.groupdict()
            body, consumed = block_at(lines, i)
            comment = leading_comment(lines, i)
            buckets[bucket].append((comment + "\n" + body).strip())

            if kind == "handler":
                signature = f"HandleEvent({groups['event']}) {groups.get('when') or 'post'}"
            elif kind == "function":
                signature = f"{groups['ret']}({' '.join(groups['args'].split())})"
            elif kind == "trigger":
                signature = f"on {groups['on']}"
            else:
                kind = groups["kind"]
                signature = kind
            decls.append({
                "module": module, "kind": kind, "name": groups["name"],
                "signature": signature, "purpose": "",
                "phase": PHASE_BY_MODULE.get(module, "?"), "status": "ACTIVE",
            })
            break

        i = consumed + 1

    return {
        "module": module,
        "phase": PHASE_BY_MODULE.get(module, "?"),
        "thesis": module_thesis(lines),
        "buckets": {k: "\n\n".join(v) for k, v in buckets.items()},
        "decls": decls,
        "source": "\n".join(lines),
        "bytes": path.stat().st_size,
    }


def load_curation() -> dict[str, dict[str, str]]:
    """Curated prose lives in its own JSON so regenerating the tab cannot lose it."""
    if CURATION.is_file():
        return json.loads(CURATION.read_text(encoding="utf-8"))
    # Bootstrap from a previously hand-maintained inventory CSV, if present.
    out: dict[str, dict[str, str]] = {}
    if INDEX_CSV.is_file():
        with INDEX_CSV.open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                key = f"{row.get('module','')}:{row.get('name','')}"
                out[key] = {
                    "purpose": row.get("purpose", ""),
                    "status": row.get("status", "ACTIVE"),
                }
    return out


def style(worksheet, header: list[str], wide: set[str]) -> None:
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, name in enumerate(header, start=1):
        letter = openpyxl.utils.get_column_letter(index)
        worksheet.column_dimensions[letter].width = 60 if name in wide else 18
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--gamedata", type=Path, default=SCEN_GAMEDATA)
    parser.add_argument("--workbook", type=Path, default=WORKBOOK)
    parser.add_argument("--check", action="store_true",
                        help="exit 1 if the tab has drifted from the code; writes nothing")
    args = parser.parse_args()

    if not args.gamedata.is_dir():
        print(f"ERROR: no such gamedata dir: {args.gamedata}", file=sys.stderr)
        return 2
    if not args.workbook.is_file():
        print(f"ERROR: no workbook at {args.workbook}; run export_mod_workbook.py first",
              file=sys.stderr)
        return 2

    order = include_order(args.gamedata / "scenario.slc")
    curated = load_curation()

    modules = [scan(p) for p in sorted(args.gamedata.glob("*.slc"))
               if p.name not in STOCK_MODULES]
    modules.sort(key=lambda m: (order.get(m["module"], 99), m["module"]))

    dim_rows: list[list[object]] = []
    idx_rows: list[list[object]] = []
    seen_keys: set[str] = set()
    missing: list[str] = []

    for mod in modules:
        key = f"{mod['module']}:-"
        seen_keys.add(key)
        purpose = curated.get(key, {}).get("purpose", "") or mod["thesis"]
        buckets = mod["buckets"]
        dim_rows.append([
            mod["module"], mod["phase"], order.get(mod["module"], 99), purpose,
            clip(buckets["constants"]), clip(buckets["functions"]),
            clip(buckets["handlers"]), clip(buckets["triggers"]),
            clip(buckets["segments"]),
            len(mod["decls"]), mod["bytes"], clip(mod["source"]),
        ])
        for decl in mod["decls"]:
            dkey = f"{decl['module']}:{decl['name']}"
            seen_keys.add(dkey)
            prior = curated.get(dkey, {})
            decl["purpose"] = prior.get("purpose", "")
            status = (prior.get("status") or "").strip().upper()
            # Never carry a stale PLANNED forward over code that demonstrably exists.
            if status and status != "PLANNED":
                decl["status"] = status
            if not decl["purpose"]:
                missing.append(dkey)
            idx_rows.append([decl[c] for c in IDX_HEADER])

    dropped = sorted(set(curated) - seen_keys)

    if args.check:
        # Excel round-trips an empty string as None, so both sides must be
        # normalised or every check reports STALE immediately after a write.
        def norm(table):
            return [["" if cell is None else str(cell) for cell in row] for row in table]

        book = openpyxl.load_workbook(args.workbook, read_only=True)
        stale = SHEET_DIM not in book.sheetnames
        if not stale:
            existing = [list(r) for r in book[SHEET_DIM].iter_rows(values_only=True)]
            stale = norm(existing) != norm([DIM_HEADER] + dim_rows)
        print(f"{SHEET_DIM} tab: {'STALE' if stale else 'current'} "
              f"({len(modules)} modules, {len(idx_rows)} declarations)")
        return 1 if stale else 0

    book = openpyxl.load_workbook(args.workbook)
    for name, header, rows, wide in (
        (SHEET_DIM, DIM_HEADER, dim_rows,
         {"purpose", "constants", "functions", "handlers", "triggers", "segments", "source"}),
        (SHEET_INDEX, IDX_HEADER, idx_rows, {"signature", "purpose"}),
    ):
        if name in book.sheetnames:
            del book[name]
        worksheet = book.create_sheet(name)
        worksheet.append(header)
        for row in rows:
            worksheet.append(row)
        style(worksheet, header, wide)
    book.save(args.workbook)

    # Persist curation so the next run cannot lose it, and mirror the flat index
    # to CSV so the tab is greppable and diffable outside Excel.
    CURATION.write_text(
        json.dumps(
            {f"{r[0]}:{r[2]}": {"purpose": r[4], "status": r[6]} for r in idx_rows}
            | {f"{r[0]}:-": {"purpose": r[3], "status": "ACTIVE"} for r in dim_rows},
            indent=2, sort_keys=True,
        ),
        encoding="utf-8",
    )
    with INDEX_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(IDX_HEADER)
        writer.writerows(idx_rows)

    print(f"wrote {SHEET_DIM} + {SHEET_INDEX} into {args.workbook.name}")
    print(f"  {len(modules)} modules, {len(idx_rows)} declarations, "
          f"{sum(m['bytes'] for m in modules)} bytes of SLIC")
    for mod in modules:
        b = mod["buckets"]
        print(f"  {mod['module']:22s} phase {mod['phase']}  "
              f"const {len(b['constants'].splitlines()):3d}  decls {len(mod['decls']):3d}")
    if dropped:
        print(f"dropped {len(dropped)} curated row(s) with no declaration in code:")
        for key in dropped:
            print(f"  {key}")
    if missing:
        print(f"{len(missing)} declaration(s) need purpose text:")
        for key in missing:
            print(f"  {key}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
