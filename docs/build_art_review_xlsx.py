"""Build an Excel workbook for human art review.

Each row = one cell from the MoMJR source sheet.
Columns: clipped image, art_cell_index, grid (row,col), current unit name, 
         vision description, blank "Correct Name" for human to fill.
"""
import csv
import json
import struct
from pathlib import Path
from io import BytesIO

import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

CSV_PATH = Path(__file__).parent.parent / "tools" / "momjr_csv" / "units.csv"
MOMJR_SHEET = Path(__file__).parent / "img" / "momjr_units_sheet.png"
AUDIT_PATH = Path(__file__).parent / "art_audit_results.json"
OUT_PATH = Path(__file__).parent.parent / "art_review.xlsx"

CELL_SIZE = 65
COLS = 9

# Background colors to remove
# Magenta: R>220, G<40, B>220
# Grey-purple: R 120-170, G 60-100, B 120-170


def clip_cell(sheet_arr, art_idx):
    """Extract a cell from the MoMJR sheet, remove background colors."""
    col = art_idx % COLS
    row = art_idx // COLS
    x0 = col * CELL_SIZE + 2
    y0 = row * CELL_SIZE + 2
    x1 = (col + 1) * CELL_SIZE - 1
    y1 = (row + 1) * CELL_SIZE - 1

    if y1 > sheet_arr.shape[0] or x1 > sheet_arr.shape[1]:
        return None, row, col

    cell = sheet_arr[y0:y1, x0:x1].copy()

    # Remove only the two exact background colors
    magenta = (cell[:, :, 0] > 220) & (cell[:, :, 1] < 40) & (cell[:, :, 2] > 220)
    grey_purple = ((cell[:, :, 0] > 120) & (cell[:, :, 0] < 170) &
                   (cell[:, :, 1] > 60) & (cell[:, :, 1] < 100) &
                   (cell[:, :, 2] > 120) & (cell[:, :, 2] < 170))
    cell[magenta] = [24, 24, 24]
    cell[grey_purple] = [24, 24, 24]

    return Image.fromarray(cell), row, col


def main():
    # Load sheet
    sheet = Image.open(MOMJR_SHEET).convert("RGB")
    sheet_arr = np.array(sheet)

    # Load audit descriptions
    descriptions = {}
    if AUDIT_PATH.exists():
        with open(AUDIT_PATH) as f:
            data = json.load(f)
        for item in data.get("mom", []):
            descriptions[item.get("slug", "")] = item.get("description", "")

    # Load unit CSV
    units = []
    with open(CSV_PATH, newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            name = r["name"].strip()
            art_idx = int(r["art_cell_index"].strip())
            slug = name.lower().replace(" ", "_").replace("'", "")
            units.append({"name": name, "art_idx": art_idx, "slug": slug})

    # Sort by art_cell_index
    units.sort(key=lambda x: x["art_idx"])

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Art Review"

    # Headers
    headers = ["Image", "art_cell_index", "Grid (r,c)", "Current Unit Name",
               "Vision Description", "Correct Name (fill this)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 25

    for idx, u in enumerate(units):
        row_num = idx + 2
        art_idx = u["art_idx"]

        # Clip image
        img, grid_r, grid_c = clip_cell(sheet_arr, art_idx)

        if img:
            # Save to buffer
            buf = BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)

            # Insert image
            xl_img = XLImage(buf)
            xl_img.width = 60
            xl_img.height = 60
            ws.add_image(xl_img, f"A{row_num}")

        # Set row height to fit image
        ws.row_dimensions[row_num].height = 50

        # Data columns
        ws.cell(row=row_num, column=2, value=art_idx)
        ws.cell(row=row_num, column=3, value=f"r{grid_r},c{grid_c}")
        ws.cell(row=row_num, column=4, value=u["name"])
        ws.cell(row=row_num, column=5, value=descriptions.get(u["slug"], ""))
        ws.cell(row=row_num, column=6, value="")  # For human to fill

    wb.save(OUT_PATH)
    print(f"Saved {OUT_PATH} with {len(units)} rows")


if __name__ == "__main__":
    main()
